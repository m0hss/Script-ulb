import openpyxl
from openpyxl.styles import Font
import time

class ExcelManager:
    def __init__(self, path):
        self.workbook = openpyxl.load_workbook(path)
        self.sheet = self.workbook.active

    def get_value(self, row, col):
        return self.sheet.cell(row, col).value

    def set_value(self, row, col, value):
        self.sheet.cell(row, col).value = value

    def find_row_by_value(self, col, value):
        for i in range(2, self.sheet.max_row + 1):
            if self.get_value(i, col) == value:
                return i
        return -1

    def delete_rows(self, start_row, end_row):
        self.sheet.delete_rows(start_row, end_row)

    def save(self, path):
        self.workbook.save(path)

class PlacementManager:
    def __init__(self, pref_path, rang_path, hopital_path, service_path, place_path):
        self.pref_sheet = ExcelManager(pref_path)
        self.rang_sheet = ExcelManager(rang_path)
        self.hopital_sheet = ExcelManager(hopital_path)
        self.service_sheet = ExcelManager(service_path)
        self.place_sheet = ExcelManager(place_path)
        self.stage = {"Matricule": [], "Hopital": [], "Service": []}

    def get_name(self, sheet_manager, val):
        row = sheet_manager.find_row_by_value(1, val)
        if row != -1:
            return sheet_manager.get_value(row, 2)
        return "Unknown"

    def find_place_by(self, hospital_id, service_id):
        for i in range(2, self.place_sheet.sheet.max_row + 1):
            if self.place_sheet.get_value(i, 4) != 0 and \
               self.place_sheet.get_value(i, 2) == hospital_id and \
               self.place_sheet.get_value(i, 3) == service_id:
                self.place_sheet.set_value(i, 4, self.place_sheet.get_value(i, 4) - 1)
                return True
        return False

    def find_general_place(self):
        for i in range(2, self.place_sheet.sheet.max_row + 1):
            if self.place_sheet.get_value(i, 4) != 0:
                self.place_sheet.set_value(i, 4, self.place_sheet.get_value(i, 4) - 1)
                return i
        return -1

    def allocate_stages(self):
        for row in range(2, self.rang_sheet.sheet.max_row + 1):
            matricule = self.rang_sheet.get_value(row, 1)
            preferences = {}

            # Gather preferences for the current matricule
            for row_pref in range(2, self.pref_sheet.sheet.max_row + 1):
                if self.pref_sheet.get_value(row_pref, 3) == matricule:
                    preference_id = self.pref_sheet.get_value(row_pref, 6)
                    preferences[preference_id] = row_pref

            preferences = dict(sorted(preferences.items()))
            flag = False

            for pref_id in preferences.values():
                hospital_id = self.pref_sheet.get_value(pref_id, 4)
                service_id = self.pref_sheet.get_value(pref_id, 5)
                preference_type = self.pref_sheet.get_value(pref_id, 7)

                if preference_type == 1 or pref_id == list(preferences.values())[-1]:
                    if self.find_place_by(hospital_id, service_id):
                        self.stage["Matricule"].append(matricule)
                        self.stage["Hopital"].append(self.get_name(self.hopital_sheet, hospital_id))
                        self.stage["Service"].append(self.get_name(self.service_sheet, service_id))
                        break

            if not preferences:
                general_place_row = self.find_general_place()
                if general_place_row != -1:
                    hospital_id = self.place_sheet.get_value(general_place_row, 2)
                    service_id = self.place_sheet.get_value(general_place_row, 3)
                    self.stage["Matricule"].append(matricule)
                    self.stage["Hopital"].append(self.get_name(self.hopital_sheet, hospital_id))
                    self.stage["Service"].append(self.get_name(self.service_sheet, service_id))

    def save_results(self, output_path):
        wb_stage = openpyxl.Workbook()
        sheet_stage = wb_stage.active

        headers = ["Matricule", "Hopital", "Service"]
        for col, header in enumerate(headers, 1):
            sheet_stage.cell(1, col, header).font = Font(size=12, bold=True)
            sheet_stage.column_dimensions[chr(64 + col)].width = 20
            sheet_stage.cell(1, col).value = header

        for i, matricule in enumerate(self.stage["Matricule"], start=2):
            sheet_stage.cell(i, 1).value = matricule
            sheet_stage.cell(i, 2).value = self.stage["Hopital"][i - 2]
            sheet_stage.cell(i, 3).value = self.stage["Service"][i - 2]

        wb_stage.save(output_path)

if __name__ == "__main__":
    start_time = time.time()
    
    placement_manager = PlacementManager(
        "./src/Annexe 1 - préférences.xlsx",
        "./src/Annexe 2 - classement.xlsx",
        "./src/Annexe 3 - hopitaux.xlsx",
        "./src/Annexe 4 - services.xlsx",
        "./src/Annexe 5 - places.xlsx"
    )

    placement_manager.allocate_stages()
    placement_manager.save_results("Stages_Affectations1.xlsx")

    print(f"--- {time.time() - start_time} seconds ---")